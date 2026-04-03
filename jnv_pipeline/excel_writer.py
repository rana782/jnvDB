from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


SHEET_HEADERS: dict[str, list[str]] = {
    "schools": [
        "udise",
        "school_name",
        "state",
        "district",
        "region_code",
        "region_name",
        "academic_year",
        "total_students",
        "total_boys",
        "total_girls",
        "source_pdf_name",
        "parse_confidence",
        "notes",
    ],
    "enrolment_social": ["udise", "category", "boys", "girls", "total"],
    "enrolment_minority": ["udise", "category", "boys", "girls", "total"],
    "enrolment_others": ["udise", "category", "boys", "girls", "total"],
    "enrolment_age": ["udise", "age_band", "boys", "girls", "total"],
    "facilities": [
        "udise",
        "water_available",
        "electricity_available",
        "internet_available",
        "solar_available",
        "playground_available",
        "library_available",
        "functional_toilets_b",
        "functional_toilets_g",
        "desktops",
        "laptops",
        "tablets",
        "printers",
        "smart_class_tv",
        "projectors",
        "medical_checkups",
        "ramps_available",
    ],
    "teachers": ["udise", "category", "label", "count"],
}


def write_workbook(sheets: dict[str, list[dict]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, headers in SHEET_HEADERS.items():
            rows = sheets.get(sheet_name, [])
            frame = pd.DataFrame(rows, columns=headers)
            frame.to_excel(writer, sheet_name=sheet_name, index=False)

    wb = load_workbook(output_path)
    for sheet_name, headers in SHEET_HEADERS.items():
        ws = wb[sheet_name]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(1, ws.max_row)}"
    wb.save(output_path)
